from __future__ import annotations

import re
from pathlib import Path
from typing import Iterable, Set

from openpyxl import Workbook

from src.models import AuditResult, DocumentArtifact, SqlAnalysis

EXPECTED_SECTIONS = {
    "1. Referencia rapida y ficha del producto",
    "2. Fuentes de datos",
    "3. Flujo general del proceso",
    "4. Matriz de trazabilidad",
    "5. Transformaciones",
    "6. Reglas de negocio",
    "7. Proceso paso a paso detallado",
    "8. Convenciones y navegacion",
}

STEP_SECTIONS = {
    "1. Fuentes de datos",
    "2. Flujo general del proceso",
    "3. Matriz de trazabilidad",
    "4. Transformaciones",
    "5. Reglas de negocio",
    "6. Proceso paso a paso detallado",
    "7. Criterio de aceptacion",
}


def audit_outputs(
    analysis: SqlAnalysis,
    document: DocumentArtifact,
    output_path: str | Path,
) -> AuditResult:
    errors = []
    warnings = []

    expected_sections = STEP_SECTIONS if document.document_kind == "step" else EXPECTED_SECTIONS
    sections = set(document.section_titles)
    if sections != expected_sections:
        missing = sorted(expected_sections - sections)
        extra = sorted(sections - expected_sections)
        if missing:
            errors.append(f"Faltan secciones obligatorias: {', '.join(missing)}")
        if extra:
            warnings.append(f"Se detectaron secciones adicionales: {', '.join(extra)}")

    final_fields = {item.field_name for item in analysis.transformations}
    section4_fields = set(document.section4_fields)
    section5_fields = set(document.section5_fields)
    missing_in_s4 = sorted(final_fields - section4_fields)
    missing_in_s5 = sorted(final_fields - section5_fields)
    if missing_in_s4:
        errors.append(f"Faltan campos finales en Seccion 4: {', '.join(missing_in_s4)}")
    if missing_in_s5:
        errors.append(f"Faltan campos finales en Seccion 5: {', '.join(missing_in_s5)}")

    defined_rules = {rule.id for rule in analysis.rules}
    referenced_rules = _referenced_rules(document.document_text) | set(document.referenced_rule_ids)
    undefined_rules = sorted(referenced_rules - defined_rules)
    if undefined_rules:
        errors.append(f"Hay reglas referenciadas que no existen en Seccion 6: {', '.join(undefined_rules)}")

    sql_markers = _sql_markers(document.document_text)
    if sql_markers:
        errors.append(
            "Se detectaron patrones de SQL literal en el documento: " + ", ".join(sorted(sql_markers))
        )

    document_name = Path(document.path).name
    if not document_name.startswith("DA_REQ_CD_IT_"):
        errors.append(
            f"El nombre del documento de salida debe iniciar con DA_REQ_CD_IT_: {document_name}"
        )

    if analysis.unresolved_variables:
        errors.append(
            "Quedaron variables sin resolver en el SQL: " + ", ".join(analysis.unresolved_variables)
        )
    elif analysis.auto_resolved_variables:
        warnings.append(
            "Se resolvieron variables con valores inferidos: "
            + ", ".join(f"{k}={v}" for k, v in sorted(analysis.auto_resolved_variables.items()))
        )

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    json_output = output.with_suffix(".json")
    result = AuditResult(
        passed=not errors,
        errors=errors,
        warnings=warnings,
        output_path=str(output),
        output_json_path=str(json_output),
    )
    _write_audit_workbook(output, result)
    json_output.write_text(result.model_dump_json(indent=2), encoding="utf-8")
    return result


def _write_audit_workbook(path: Path, result: AuditResult) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    ws.append(["Estado", "Valor"])
    ws.append(["passed", "SI" if result.passed else "NO"])
    ws.append(["output_path", result.output_path or ""])
    ws.append(["output_json_path", result.output_json_path or ""])

    ws_errors = wb.create_sheet("Errores")
    ws_errors.append(["#", "Error"])
    if result.errors:
        for index, error in enumerate(result.errors, start=1):
            ws_errors.append([index, error])
    else:
        ws_errors.append([1, "Sin errores"])

    ws_warnings = wb.create_sheet("Warnings")
    ws_warnings.append(["#", "Warning"])
    if result.warnings:
        for index, warning in enumerate(result.warnings, start=1):
            ws_warnings.append([index, warning])
    else:
        ws_warnings.append([1, "Sin warnings"])

    wb.save(path)


def _referenced_rules(text: str) -> Set[str]:
    return set(re.findall(r"\bRN-\d{2}\b", text))


def _sql_markers(text: str) -> Set[str]:
    patterns = {
        "select from": r"\bselect\b.+\bfrom\b",
        "insert": r"\binsert\b",
        "join on": r"\bjoin\b.+\bon\b",
        "insert overwrite": r"\binsert\s+overwrite\b",
        "compute stats": r"\bcompute\s+stats\b",
        "variable sin resolver": r"\$\{[^}]+\}",
    }
    found = set()
    compact = " ".join(text.lower().split())
    for label, pattern in patterns.items():
        if re.search(pattern, compact):
            found.add(label)
    return found
